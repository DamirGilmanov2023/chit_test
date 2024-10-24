from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
import urllib.request
import time
import random
from openpyxl import load_workbook

def input_class(class_,keys,i,time_before,time_after):
    global driver
    time.sleep(time_before)
    iter_error=0
    while driver.find_elements(By.CLASS_NAME,class_)[i].get_attribute('value')!=keys:
        time.sleep(1)
        try:
            driver.find_elements(By.CLASS_NAME,class_)[i].clear()
            time.sleep(1)
            driver.find_elements(By.CLASS_NAME,class_)[i].send_keys(keys)
        except:
            pass
        iter_error+=1
        if iter_error>5:
            break
    time.sleep(time_after)

def click_one_click(class_,name_class,i,time_before,time_after):
    global driver
    time.sleep(time_before)
    flag=True
    iter_error = 0
    while flag:
        try:
            btn = driver.find_elements(By.CLASS_NAME,class_)[i]
            assert btn.text == name_class
            btn.click()
            print(f"Клик по {name_class}")
            flag=False
        except:
            time.sleep(1)
            iter_error += 1
            if iter_error >= 3:
                try:
                    btn.click()
                    print(f"Клик по {name_class}")
                except:
                    pass
                flag = False
    time.sleep(time_after)

def click_tagname_by_number(tagname,number,time_before,time_after):
    global driver
    time.sleep(time_before)
    flag=True
    iter_error = 0
    while flag:
        try:
            tag = driver.find_elements(By.TAG_NAME,tagname)[number]
            tag.click()
            flag=False
        except:
            time.sleep(1)
            iter_error += 1
            if iter_error >= 5:
                flag = False
    time.sleep(time_after)

def click_id_noname(id_,time_before,time_after):
    global driver
    time.sleep(time_before)
    flag=True
    iter_error = 0
    while flag:
        try:
            btn = driver.find_element(By.ID,id_)
            btn.click()
            print(f"Клик по id {id_}")
            flag=False
        except:
            time.sleep(1)
            iter_error += 1
            if iter_error >= 5:
                try:
                    btn.click()
                    print(f"Клик по id {id_}")
                except:
                    pass
                flag = False
    time.sleep(time_after)

#Открытие файла с логинами и паролями
auth=[]
auth_iterator=0
with open('auth.txt','r') as file:
    for line in file:
        l=line.replace(" ","")
        if l!="":
            lr=[]
            lr=l.split(":")
            auth.append([lr[0],lr[1]])

#Ввод ссылки с тестом
url=input("Введите ссылку на тест:")

#Ввод количества циклов
repeat=input("Введите количество циклов:")
rep=0

while rep<int(repeat):
	try:
		wb=load_workbook("./output.xlsx")
		true=[]
		for i in range(1,wb["true"].max_row):
		    vrm=[]
		    vrm.append(wb["true"][f"A{i+1}"].value)#Заголовок
		    vrm.append(wb["true"][f"B{i+1}"].value)#Тип
		    vrm.append(wb["true"][f"D{i+1}"].value)#Изображение
		    vrm.append(wb["true"][f"E{i+1}"].value)#Правильный вариант
		    vrm_m=[]
		    n=6
		    while 1:
		        cell=wb["true"].cell(row=i+1, column=n).value
		        if cell!=None:
		            vrm_m.append(cell)#варианты
		        else:
		            break
		        n+=1
		    vrm.append(vrm_m)
		    true.append(vrm)
		false=[]
		for i in range(1,wb["not"].max_row):
		    vrm=[]
		    vrm.append(wb["not"][f"A{i+1}"].value)#Заголовок
		    vrm.append(wb["not"][f"B{i+1}"].value)#Тип
		    vrm.append(wb["not"][f"D{i+1}"].value)#Изображение
		    vrm.append(wb["not"][f"E{i+1}"].value)#Не правильный вариант
		    vrm_m=[]
		    n=6
		    while 1:
		        cell=wb["not"].cell(row=i+1, column=n).value
		        if cell!=None:
		            vrm_m.append(cell)#варианты
		        else:
		            break
		        n+=1
		    vrm.append(vrm_m)
		    false.append(vrm)
		options=Options()
		firefox_profile = FirefoxProfile()
		options.profile = firefox_profile
		driver = webdriver.Firefox(options=options)
		#driver=webdriver.Firefox(executable_path="./geckodriver")
		driver.set_page_load_timeout(30)
		driver.implicitly_wait(2)
		driver.get(url) 

		input_class("input",auth[auth_iterator][0],0,2,2)
		input_class("input",auth[auth_iterator][1],1,2,2)
		click_one_click('is-primary','Войти',0,2,12)
		driver.switch_to.frame("aos-frame")
		click_tagname_by_number('input',5,2,2)

		step1=True
		while step1:
			time.sleep(8)
			vopros_content=driver.find_element(By.ID,"dfdTestWinMain_2")
			Zag=vopros_content.find_elements(By.TAG_NAME,"span")[0]
			zag_text=Zag.text
			try:
				Img=vopros_content.find_elements(By.TAG_NAME,"img")[0]
				attr_img=Img.get_attribute("src")
				spl_attr_img=attr_img.split("/")
				name_img=spl_attr_img[len(spl_attr_img)-1]
				urllib.request.urlretrieve(attr_img,f"./img/{name_img}")
			except:
				attr_img=""
				name_img=""
			table=vopros_content.find_elements(By.TAG_NAME,"table")[0]
			input_=table.find_elements(By.TAG_NAME,"input")
			input_type=input_[0].get_attribute("type")
			label=table.find_elements(By.TAG_NAME,"label")
			label_mass=[]
			for lbl in label:
				lll=lbl.text
				ll=lll.split(")       ")
				label_mass.append(ll[1])
			#print(label[0].text)
			#print(label[1].text)
			#print(input_type)
			flag_next=False
			for tru in true:
				if zag_text==tru[0] and \
				input_type==tru[1] and \
				name_img==tru[2]:
					kol_vo=0
					for lbl in label_mass:
						if lbl in tru[4]:
							kol_vo+=1
					if kol_vo==len(label_mass):
						flag_next=True
			if flag_next:
				click_id_noname("dfdTestButton_Next",2,2)
			else:
				step1=False
		#step2
		not_variant=[]
		for fals in false:
		    if zag_text==fals[0] and \
		    input_type==fals[1] and \
		    name_img==fals[2]:
		        kol_vo=0
		        for lbl in label_mass:
		            if lbl in fals[4]:
		                kol_vo+=1
		        if kol_vo==len(label_mass):
		            not_variant.append(fals[3])
		#step3
		#body = driver.find_elements(By.TAG_NAME,'body')[0]
		#body.send_keys(Keys.PAGE_DOWN)
		try:
			vopros_content.send_keys(Keys.PAGE_DOWN)
			vopros_content.send_keys(Keys.END)
		except:
			pass
		if input_type=="checkbox":
		    flag_step3=True
		    while flag_step3:
		        check=random.randint(0,len(input_)-1)
		        not_check=check
		        if label_mass[not_check] not in not_variant:
		            i=0
		            while i<len(input_):
		                if i!=not_check:
		                    input_[i].click()
		                i+=1
		            flag_step3=False
		elif input_type=="radio":
		    flag_step3=True
		    while flag_step3:
		        check=random.randint(0,len(input_)-1)
		        if label_mass[check] not in not_variant:
		            input_[check].click()
		            flag_step3=False
		#step4
		target = driver.find_element(By.ID,"dfdTestButton_Next")
		driver.execute_script('arguments[0].scrollIntoView(true);', target)
		click_id_noname("dfdTestButton_Next",2,2)
		elem=driver.find_element(By.ID,"dfdTestButton_Abort")#.set_attribyte("disabled","false")
		driver.execute_script("arguments[0].removeAttribute('disabled')",elem)
		click_id_noname("dfdTestButton_Abort",2,2)
		driver.find_elements(By.CLASS_NAME,"dfdButton")[0].click()
		time.sleep(2)

		rez_table=driver.find_elements(By.TAG_NAME,"table")[2]
		#print(rez_table.text)
		rez_=rez_table.find_elements(By.TAG_NAME,"td")[3].text
		rez_=rez_.split(" ")
		rez=rez_[0]
		#print(rez)
		vrm_mass=[]
		#print(zag_text)
		vrm_mass.append(zag_text)
		vrm_mass.append(input_type)
		vrm_mass.append(attr_img)
		vrm_mass.append(name_img)
		vrm_mass.append(label_mass[check])
		for lbl in label_mass:
		    vrm_mass.append(lbl)
		if int(rez)==0:
		    wb["not"].append(vrm_mass)
		elif int(rez)==1:
		    wb["true"].append(vrm_mass)
		wb.save("./output.xlsx")
		driver.quit()
		auth_iterator+=1
		if auth_iterator>=len(auth):
			auth_iterator=0
		rep+=1
	except:
		print("Ошибка")
		try:
			driver.quit()
		except:
			pass