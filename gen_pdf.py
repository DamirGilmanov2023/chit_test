from openpyxl import load_workbook
#from weasyprint import HTML

wb=load_workbook("./output.xlsx")
true=[]
for i in range(1,wb["true"].max_row):
    vrm=[]
    vrm.append(wb["true"][f"A{i+1}"].value)#Заголовок
    vrm.append(wb["true"][f"B{i+1}"].value)#Тип
    vrm.append(wb["true"][f"D{i+1}"].value)#Изображение
    vrm.append(wb["true"][f"E{i+1}"].value)#Правильный вариант
    vrm_m=[]
    n=6
    while 1:
        cell=wb["true"].cell(row=i+1, column=n).value
        if cell!=None:
            vrm_m.append(cell)#варианты
        else:
            break
        n+=1
    vrm.append(vrm_m)
    true.append(vrm)
html=""
for tr in true:
	html+=f"<h3>{tr[0]}</h3>\n"
	if tr[2]!=None:
		html+=f"<p><img src='./img/{tr[2]}'></p>\n"
	if tr[1]=="checkbox":
		for c in tr[4]:
			if c==tr[3]:
				html+=f"<p style=\"color:red\">{c}</p>\n"
			else:
				html+=f"<p style=\"color:green\">{c}</p>\n"
	elif tr[1]=="radio":
		for r in tr[4]:
			if r==tr[3]:
				html+=f"<p style=\"color:green\">{r}</p>\n"
			else:
				html+=f"<p style=\"color:red\">{r}</p>\n"

#HTML(string=html,base_url="base_url").write_pdf('out.pdf')

from pdfrw import PdfWriter
y = PdfWriter()
y.addpage(html)
y.write('result.pdf')
