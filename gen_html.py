from openpyxl import load_workbook

wb=load_workbook("./output.xlsx")
true=[]
for i in range(1,wb["true"].max_row):
    vrm=[]
    vrm.append(wb["true"][f"A{i+1}"].value)#Заголовок
    vrm.append(wb["true"][f"B{i+1}"].value)#Тип
    vrm.append(wb["true"][f"D{i+1}"].value)#Изображение
    vrm.append(wb["true"][f"E{i+1}"].value)#Правильный вариант
    vrm_m=[]
    n=6
    while 1:
        cell=wb["true"].cell(row=i+1, column=n).value
        if cell!=None:
            vrm_m.append(cell)#варианты
        else:
            break
        n+=1
    vrm.append(vrm_m)
    true.append(vrm)
with open("out.html","w") as file:
	for tr in true:
		file.write(f"<h3>{tr[0]}</h3>")
		if tr[2]!=None:
			file.write(f"<p><img src='./img/{tr[2]}'></p>")
		if tr[1]=="checkbox":
			for c in tr[4]:
				if c==tr[3]:
					file.write(f"<p><input type='checkbox' disabled>{c}</p>")
				else:
					file.write(f"<p><input type='checkbox' checked disabled>{c}</p>")
		elif tr[1]=="radio":
			for r in tr[4]:
				if r==tr[3]:
					file.write(f"<p><input type='radio' checked disabled>{r}</p>")
				else:
					file.write(f"<p><input type='radio' disabled>{r}</p>")
